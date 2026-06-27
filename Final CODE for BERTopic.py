import os
import pandas as pd
import numpy as np
import random
import torch
from bertopic import BERTopic
from sentence_transformers import SentenceTransformer
from umap import UMAP
from hdbscan import HDBSCAN
from gensim.models.coherencemodel import CoherenceModel
from gensim.corpora import Dictionary
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def main():
    # Global Configuration: Set Seeds for Reproducibility
    SEED = 42
    np.random.seed(SEED)
    random.seed(SEED)
    torch.manual_seed(SEED)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(SEED)

    print("Loading population dataset for BERTopic...")
    # Load your real saved preprocessed data
    df = pd.read_pickle("preprocessed_topic_population.pkl")

    # CRITICAL: BERTopic needs the original raw string text headlines to analyze full sentence context
    docs = df['headlines'].tolist()
    tokenized_docs = df['tokens_traditional'].tolist()
    TOTAL_DOCS = len(docs)

    print("Initializing Transformer Pipeline components...")
    embedding_model = SentenceTransformer("all-MiniLM-L6-v2")
    umap_model = UMAP(n_neighbors=15, n_components=5, min_dist=0.0, random_state=42)
    hdbscan_model = HDBSCAN(min_cluster_size=15, metric='euclidean', cluster_selection_method='eom',
                            prediction_data=True)

    print("Training BERTopic (Auto-detecting optimal topic clusters)...")
    topic_model = BERTopic(
        embedding_model=embedding_model,
        umap_model=umap_model,
        hdbscan_model=hdbscan_model,
        calculate_probabilities=True,
        verbose=True
    )

    # Fit the model on your raw headlines
    topics, probs = topic_model.fit_transform(docs)
    df['Topic'] = topics

    topic_info = topic_model.get_topic_info()
    valid_topics = topic_info[topic_info['Topic'] != -1].copy()
    NUM_TOPICS = len(valid_topics)
    print(f"\nSuccess! BERTopic organically auto-detected {NUM_TOPICS} distinct topics (excluding noise).")

    print("\nComputing BERTopic Cv Coherence score using the evaluation framework...")
    all_topics = topic_model.get_topics()
    bertopic_words = []

    # Sort topics cleanly to populate the table chronologically
    sorted_topic_ids = sorted([t for t in all_topics.keys() if t != -1])

    for topic_id in sorted_topic_ids:
        # Extract top 8 words per topic to match LDA/NMF format parameters perfectly
        words = [word for word, _ in all_topics[topic_id][:8]]
        bertopic_words.append(words)

    gensim_dictionary = Dictionary(tokenized_docs)
    coherence_model = CoherenceModel(
        topics=bertopic_words,
        texts=tokenized_docs,
        dictionary=gensim_dictionary,
        coherence='c_v'
    )

    overall_score = coherence_model.get_coherence()
    individual_scores = coherence_model.get_coherence_per_topic()

    # =========================================================================
    # SECTION 1: DISPLAY THE TABLE IN THE CONSOLE
    # =========================================================================
    print("\n" + "=" * 85)
    print(" BERTopic MODEL RESULT SUMMARY (MATCHING LDA/NMF LAYOUT)")
    print("=" * 85)
    print(f"Auto-Detected Topic Count: {NUM_TOPICS}")
    print(f"Overall BERTopic Cv Coherence Score: {overall_score:.4f}")
    print("-" * 85)
    print(
        f"| {'Topic':<5} | {'Thematic Label':<20} | {'Dominant Keywords (Top 8)':<30} | {'Count':<7} | {'Cv Score':<8} |")
    print("|" + "-" * 7 + "|" + "-" * 22 + "|" + "-" * 32 + "|" + "-" * 9 + "|" + "-" * 10 + "|")

    for i, topic_id in enumerate(sorted_topic_ids):
        cv_score = individual_scores[i]
        kw_str = ", ".join(bertopic_words[i][:5])  # Short preview for console
        print(
            f"| {topic_id + 1:<5} | {'ENTER LABEL HERE':<20} | {kw_str:<30} | {valid_topics.iloc[i]['Count']:<7,} | {cv_score:<8.4f} |")

    print("|" + "-" * 7 + "|" + "-" * 22 + "|" + "-" * 32 + "|" + "-" * 9 + "|" + "-" * 10 + "|")
    noise_count = topic_info[topic_info['Topic'] == -1]['Count'].values[0] if -1 in topic_info['Topic'].values else 0
    print(f"Note: Noise Cluster (Topic -1) caught {noise_count} unclassifiable headlines.")
    print("=" * 85)

    # =========================================================================
    # SECTION 2: CREATE AND SAVE THE WORKBOOK EXCEL FILE AUTOMATICALLY
    # =========================================================================
    print("\nGenerating formatted Excel Worksheet matrix...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BERTopic Topics"

    # Exact Style Layout Palette from LDA/NMF script
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    noise_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB",
                             fill_type="solid")  # Light color for noise row
    summary_fill = PatternFill(start_color="E6EDF5", end_color="E6EDF5", fill_type="solid")

    thin_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    summary_border = Border(left=thin_side, right=thin_side, top=thin_side,
                            bottom=Side(border_style="double", color="111827"))

    # Title Blocks matching LDA/NMF template setup exactly
    ws.append(["Real BERTopic Topic Matrix"])
    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

    ws.append([f"Generated via Evaluation Framework (K = {NUM_TOPICS})"])
    ws.merge_cells("A2:E2")
    ws.append([])  # Spacer row

    # Correct Column Order Arrangement Matching traditional templates
    headers = ["Topic Number", "Thematic Label (Write in Word)", "Dominant Keywords", "Document Count",
               "Individual Cv Score"]
    ws.append(headers)
    ws.row_dimensions[4].height = 25

    for c in range(1, 6):
        cell = ws.cell(row=4, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 1. Populate valid organic topics first (Topics 1 to 17)
    for idx, topic_id in enumerate(sorted_topic_ids):
        row_num = idx + 5
        doc_count = valid_topics.iloc[idx]['Count']
        cv_score = individual_scores[idx]
        keywords_str = ", ".join(bertopic_words[idx])

        ws.append([idx + 1, "ENTER LABEL HERE", keywords_str, doc_count, cv_score])
        ws.row_dimensions[row_num].height = 20
        row_fill = zebra_fill if row_num % 2 == 0 else white_fill

        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=3).font = Font(name="Consolas", size=10)
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4).number_format = "#,##0"
        ws.cell(row=row_num, column=5).number_format = "0.00"
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")

        for c in range(1, 6):
            data_cell = ws.cell(row=row_num, column=c)
            data_cell.fill = row_fill
            data_cell.border = thin_border

    # 2. INSERTING THE NOISE ROW (TOPIC -1) TO MAKE IT VISIBLE
    noise_row_idx = NUM_TOPICS + 5
    noise_count = topic_info[topic_info['Topic'] == -1]['Count'].values[0] if -1 in topic_info['Topic'].values else 0

    # Using "N/A" in Cv Score column to protect math calculation
    ws.append(["Topic -1", "Outlier Noise Cluster", "Unclassifiable regional headlines and sparse vocabulary tokens",
               noise_count, "N/A"])
    ws.row_dimensions[noise_row_idx].height = 20

    ws.cell(row=noise_row_idx, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=noise_row_idx, column=1).font = Font(name="Calibri", size=11, italic=True)
    ws.cell(row=noise_row_idx, column=2).font = Font(name="Calibri", size=11, italic=True, color="595959")
    ws.cell(row=noise_row_idx, column=3).font = Font(name="Consolas", size=10, italic=True, color="595959")
    ws.cell(row=noise_row_idx, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=noise_row_idx, column=4).number_format = "#,##0"
    ws.cell(row=noise_row_idx, column=5).alignment = Alignment(horizontal="center")
    ws.cell(row=noise_row_idx, column=5).font = Font(name="Calibri", size=11, italic=True, color="7F7F7F")

    for c in range(1, 6):
        noise_cell = ws.cell(row=noise_row_idx, column=c)
        noise_cell.fill = noise_fill
        noise_cell.border = thin_border

    # 3. Summary Bottom Row Configuration
    tot_idx = noise_row_idx + 1
    ws.append(
        ["---", "OVERALL MODEL AVERAGE SCORE", "Arithmetic mean score (Valid Clusters)", TOTAL_DOCS, overall_score])
    ws.row_dimensions[tot_idx].height = 22

    ws.cell(row=tot_idx, column=4).number_format = "#,##0"
    ws.cell(row=tot_idx, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=tot_idx, column=4).font = Font(bold=True)
    ws.cell(row=tot_idx, column=5).number_format = "0.00"
    ws.cell(row=tot_idx, column=5).alignment = Alignment(horizontal="center")
    ws.cell(row=tot_idx, column=5).font = Font(bold=True)

    for c in range(1, 6):
        summary_cell = ws.cell(row=tot_idx, column=c)
        summary_cell.fill = summary_fill
        summary_cell.border = summary_border

    # Standardized template column width allocations
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20

    excel_filename = "BERTopic Matrix Final.xlsx"

    try:
        wb.save(excel_filename)
        print(
            f"\nSuccess! Your aligned BERTopic spreadsheet matrix with VISIBLE Noise row saved as '{excel_filename}'")
    except PermissionError:
        print(f"\nError: Please close '{excel_filename}' in Microsoft Excel before running the script!")

if __name__ == '__main__':
    main()