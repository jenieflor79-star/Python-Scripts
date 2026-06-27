import os
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from gensim.corpora import Dictionary
from gensim.models import LdaModel, Nmf
from gensim.models.coherencemodel import CoherenceModel


def get_dominant_topic_counts(model, corpus, num_topics):
    """
    Calculates how many documents have a specific topic as the dominant one.
    """
    topic_counts = {i: 0 for i in range(num_topics)}

    for bow in corpus:
        # Get the topic distribution for a single document
        topic_distribution = model.get_document_topics(bow) if hasattr(model, 'get_document_topics') else model[bow]

        if topic_distribution:
            # Find the topic with the highest weight/probability
            dominant_topic = max(topic_distribution, key=lambda x: x[1])[0]
            topic_counts[dominant_topic] += 1

    return topic_counts


def main():
    # 1. Load your real data and rebuild your exact models
    print("Loading preprocessed data...")
    df = pd.read_pickle("preprocessed_topic_population.pkl")
    processed_docs = df['tokens_traditional'].tolist()
    dictionary = Dictionary(processed_docs)
    dictionary.filter_extremes(no_below=2, no_above=0.5)
    corpus = [dictionary.doc2bow(doc) for doc in processed_docs]

    NUM_TOPICS = 10
    TOTAL_DOCS = len(processed_docs)

    print("Training your real models...")
    lda_model = LdaModel(corpus=corpus, id2word=dictionary, num_topics=NUM_TOPICS, random_state=42, passes=10)
    nmf_model = Nmf(corpus=corpus, id2word=dictionary, num_topics=NUM_TOPICS, random_state=42, passes=10)

    # 2. Calculate your real scores
    print("Computing real coherence values...")
    lda_coherence = CoherenceModel(model=lda_model, texts=processed_docs, dictionary=dictionary, coherence='c_v')
    nmf_coherence = CoherenceModel(model=nmf_model, texts=processed_docs, dictionary=dictionary, coherence='c_v')

    lda_overall = lda_coherence.get_coherence()
    nmf_overall = nmf_coherence.get_coherence()

    lda_individuals = lda_coherence.get_coherence_per_topic()
    nmf_individuals = nmf_coherence.get_coherence_per_topic()

    # 3. Calculate Document Counts per Topic
    print("Counting documents per topic...")
    lda_counts = get_dominant_topic_counts(lda_model, corpus, NUM_TOPICS)
    nmf_counts = get_dominant_topic_counts(nmf_model, corpus, NUM_TOPICS)

    # 4. Setup Excel Export
    wb = openpyxl.Workbook()
    ws_lda = wb.active
    ws_lda.title = "LDA Topics"
    ws_nmf = wb.create_sheet(title="NMF Topics")

    # Formatting styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    summary_fill = PatternFill(start_color="E6EDF5", end_color="E6EDF5", fill_type="solid")

    thin_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    summary_border = Border(left=thin_side, right=thin_side, top=thin_side,
                            bottom=Side(border_style="double", color="111827"))

    def build_sheet(ws, title_text, model, individual_scores, overall_score, topic_counts):
        ws.append([title_text])
        ws.merge_cells("A1:E1")
        ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
        ws.append(["Generated via Evaluation Framework (K = 10)"])
        ws.merge_cells("A2:E2")
        ws.append([])  # Spacer

        # Added "Document Count" to Column D, moved Cv Score to Column E
        ws.append(["Topic Number", "Thematic Label (Write in Word)", "Dominant Keywords", "Document Count",
                   "Individual Cv Score"])
        ws.row_dimensions[4].height = 25
        for c in range(1, 6):
            ws.cell(row=4, column=c).fill = header_fill
            ws.cell(row=4, column=c).font = header_font
            ws.cell(row=4, column=c).alignment = Alignment(horizontal="center", vertical="center")

        # Extract keywords directly out of your actual model
        for idx in range(NUM_TOPICS):
            r_words = [word for word, prop in model.show_topic(idx, topn=8)]
            keywords_str = ", ".join(r_words)
            row_num = idx + 5

            # Placing document counts per topic
            ws.append([idx + 1, "ENTER LABEL HERE", keywords_str, topic_counts[idx], individual_scores[idx]])
            ws.row_dimensions[row_num].height = 20
            row_fill = zebra_fill if row_num % 2 == 0 else white_fill

            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=3).font = Font(name="Consolas", size=10)
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=4).number_format = "#,##0"  # Format for integer counts
            ws.cell(row=row_num, column=5).number_format = "0.00"
            ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")

            for c in range(1, 6):
                ws.cell(row=row_num, column=c).fill = row_fill
                ws.cell(row=row_num, column=c).border = thin_border

        # Total Row
        tot_idx = NUM_TOPICS + 5
        # Displaying total count of all documents under the Document Count column
        ws.append(["---", "OVERALL MODEL AVERAGE SCORE", "Arithmetic mean score", TOTAL_DOCS, overall_score])
        ws.cell(row=tot_idx, column=4).number_format = "#,##0"
        ws.cell(row=tot_idx, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=tot_idx, column=4).font = Font(bold=True)
        ws.cell(row=tot_idx, column=5).number_format = "0.00"
        ws.cell(row=tot_idx, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=tot_idx, column=5).font = Font(bold=True)

        for c in range(1, 6):
            ws.cell(row=tot_idx, column=c).fill = summary_fill
            ws.cell(row=tot_idx, column=c).border = summary_border

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 65
        ws.column_dimensions['D'].width = 18  # Width for Document Count
        ws.column_dimensions['E'].width = 20  # Width for Cv Score

    build_sheet(ws_lda, "Real LDA Topic Matrix", lda_model, lda_individuals, lda_overall, lda_counts)
    build_sheet(ws_nmf, "Real NMF Topic Matrix", nmf_model, nmf_individuals, nmf_overall, nmf_counts)

    output_filename = "LDA VS NMF Matrix 3.xlsx"
    wb.save(output_filename)
    print(f"Success! Your REAL numbers with Document Counts are saved in '{output_filename}'")


if __name__ == '__main__':
    main()
