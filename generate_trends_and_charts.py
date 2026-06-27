import os
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def main():
    # =========================================================================
    # 📦 STAGE 1: DATA INGESTION & CHRONOLOGICAL PARSING
    # =========================================================================
    project_dir = r"C:\Users\4Ps BARMM\Desktop\PythonProject3"
    input_file = os.path.join(project_dir, "BARMM_Full_Population_Sentiment.xlsx")
    excel_output = os.path.join(project_dir, "BARMM_Final_Statistical_Tables.xlsx")

    print("📦 Loading the processed population dataset...")
    if not os.path.exists(input_file):
        print(f"❌ Error: The file could not be found at {input_file}.")
        return

    df = pd.read_excel(input_file)
    df.columns = df.columns.str.strip()

    print("📅 Parsing chronological vectors and filtering null properties...")
    if 'date' not in df.columns:
        print("❌ Error: The column 'date' does not exist in your Excel file.")
        return

    # 1. Linisin ang string mula sa anumang spaces
    df['date'] = df['date'].astype(str).str.strip()

    # 2. I-convert gamit ang format na gitling (%d-%m-%Y)
    # Ang errors='coerce' ay gagawing NaT (Not a Time) ang mga maling date format
    df['Clean_Date'] = pd.to_datetime(df['date'], format='%d-%m-%Y', errors='coerce')

    # DEBUG: Tingnan kung may hindi na-parse (dapat ay 0)
    invalid_dates = df['Clean_Date'].isna().sum()
    print(f"DEBUG: Rows na hindi na-parse ang date: {invalid_dates}")

    # I-extract ang taon
    df['Year'] = df['Clean_Date'].dt.year

    # 3. Filter rows (i-drop ang mga rows na walang year o sentiment)
    df = df.dropna(subset=['Year', 'RoBERTa_Full_Sentiment'])
    df['Year'] = df['Year'].astype(int)

    total_population = len(df)

    # Safety Check: Kung walang data, huminto agad
    if total_population == 0:
        print("❌ CRITICAL ERROR: Walang natirang data matapos ang filtering.")
        return

    print(f"📋 Total Rows available after structural filtering: {total_population}")
    # ... pagkatapos ng df['Year'] = df['Clean_Date'].dt.year ...

    # DEBUG: I-print natin ang data bago i-graph para makita ang values
    debug_df = df.groupby(['Year', 'RoBERTa_Full_Sentiment']).size().unstack(fill_value=0)
    debug_pct = debug_df.div(debug_df.sum(axis=1), axis=0) * 100
    print("DEBUG: Percentages per Year:")
    print(debug_pct)

    # ... ituloy ang natitirang code ...
    total_population = len(df)

    # 4. DAGDAG NA SAFETY CHECK:
    if total_population == 0:
        print("❌ CRITICAL ERROR: Walang natirang data matapos ang filtering. Pakisuri ang date format sa Excel.")
        return

    # =========================================================================
    # 📊 STAGE 2: MATPLOTLIB & SEABORN DATA VISUALIZATION GENERATION
    # =========================================================================
    print("\n📈 Generating statistical charts...")

    # Part 2.1: Overall Sentiment Distribution Chart
    sentiment_counts = df['RoBERTa_Full_Sentiment'].value_counts()
    plt.figure(figsize=(8, 5.5))
    sns.set_theme(style="whitegrid")

    colors_dist = {'Positive': '#4F81BD', 'Neutral': '#7F7F7F', 'Negative': '#C00000'}
    ax1 = sns.countplot(x='RoBERTa_Full_Sentiment', data=df, hue='RoBERTa_Full_Sentiment',
                        palette=colors_dist, order=['Negative', 'Neutral', 'Positive'], legend=False)

    for p in ax1.patches:
        height = p.get_height()
        pct = (height / total_population) * 100
        ax1.annotate(f"{int(height)}\n({pct:.2f}%)",
                     (p.get_x() + p.get_width() / 2., height),
                     ha='center', va='center', xytext=(0, 12), textcoords='offset points',
                     fontsize=10, fontweight='bold')

    plt.title('Overall Sentiment Distribution of BARMM News Headlines (2021-2025)', fontsize=12, fontweight='bold',
              pad=20)
    plt.xlabel('Sentiment Frame', fontsize=11, fontweight='bold')
    plt.ylabel('Number of Headlines (Frequency)', fontsize=11, fontweight='bold')
    plt.ylim(0, max(sentiment_counts) + (max(sentiment_counts) * 0.15))
    plt.tight_layout()

    dist_chart_path = os.path.join(project_dir, "overall_sentiment_distribution.png")
    plt.savefig(dist_chart_path, dpi=300)
    plt.close()
    print(f"🖼️ Saved Chart: {dist_chart_path}")

    # Part 2.2: Chronological Longitudinal Media Framing Trend Chart
    trend_df = df.groupby(['Year', 'RoBERTa_Full_Sentiment']).size().unstack(fill_value=0)
    trend_pct = trend_df.div(trend_df.sum(axis=1), axis=0) * 100

    plt.figure(figsize=(10, 6))

    for sentiment, marker, color in [('Positive', 'o', '#4F81BD'), ('Neutral', 's', '#7F7F7F'),
                                     ('Negative', '^', '#C00000')]:
        if sentiment in trend_pct.columns:
            plt.plot(trend_pct.index, trend_pct[sentiment], marker=marker,
                     linewidth=2.5, color=color, label=sentiment)

            for x, y in zip(trend_pct.index, trend_pct[sentiment]):
                # Explicit label positioning
                if x == 2025:
                    if sentiment == 'Positive':
                        label_y = y - 4  # Positive sa ilalim
                    elif sentiment == 'Negative':
                        label_y = y + 4  # Negative sa ibabaw
                    else:
                        label_y = y + 4
                else:
                    # Default para sa 2021-2024
                    offsets = {'Positive': -8, 'Neutral': 8, 'Negative': 8}
                    label_y = y + offsets.get(sentiment, 0)

                plt.annotate(f'{y:.1f}%', (x, y), xytext=(0, (label_y - y) * 2),
                             textcoords='offset points', ha='center',
                             fontsize=9, fontweight='bold', color=color)

    plt.title('Chronological Media Framing Trend of BARMM News Headlines (2021-2025)', fontsize=12, fontweight='bold',
              pad=20)
    plt.xlabel('Publication Year', fontsize=11, fontweight='bold')
    plt.ylabel('Proportion of Headlines per Year (%)', fontsize=11, fontweight='bold')
    plt.xticks(trend_pct.index)
    plt.ylim(0, 110)  # Sapat na space para sa mga labels
    plt.grid(True, linestyle='--', alpha=0.5)
    plt.legend(title="Sentiment Classes", loc='upper right')
    plt.tight_layout()

    trend_chart_path = os.path.join(project_dir, "chronological_media_framing_trend.png")
    plt.savefig(trend_chart_path, dpi=300)
    plt.close()
    print(f"🖼️ Saved Chart: {trend_chart_path}")
    # =========================================================================
    # 💎 STAGE 3: OPENPYXL FORMATTED MATRIX COMPILATION
    # =========================================================================
    print("\nExcel Reporting Layer Compilation...")
    wb = Workbook()

    # Establish academic styling templates
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    body_font = Font(name=font_family, size=11, bold=False, color="000000")
    total_font = Font(name=font_family, size=11, bold=True, color="000000")

    header_fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
    total_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    double_border_side = Side(border_style="double", color="000000")
    thin_top_side = Side(border_style="thin", color="000000")

    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    total_border = Border(top=thin_top_side, bottom=double_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # --- Sheet 1 Setup ---
    ws1 = wb.active
    ws1.title = "Overall Distribution"
    ws1.views.sheetView[0].showGridLines = True

    ws1.append([])
    ws1.append(["Table 1. Overall Sentiment Distribution of BARMM Headlines (2021-2025)"])
    ws1.cell(row=2, column=1).font = Font(name=font_family, size=12, bold=True)
    ws1.append([])

    ws1.append(["Sentiment Frame", "Frequency (n)", "Relative Proportion (%)"])
    for sentiment in ['Neutral', 'Negative', 'Positive']:
        count = int(df['RoBERTa_Full_Sentiment'].value_counts().get(sentiment, 0))
        ws1.append([sentiment, count, count / total_population])

    ws1.append(["Total", f"=SUM(B5:B7)", f"=SUM(C5:C7)"])

    # Format Sheet 1 Styles
    for col in range(1, 4):
        cell = ws1.cell(row=4, column=col)
        cell.font, cell.fill, cell.alignment = header_font, header_fill, (align_center if col > 1 else align_left)

    for r_idx in range(5, 8):
        ws1.cell(row=r_idx, column=1).alignment, ws1.cell(row=r_idx, column=1).font, ws1.cell(row=r_idx,
                                                                                              column=1).border = align_left, body_font, cell_border
        ws1.cell(row=r_idx, column=2).alignment, ws1.cell(row=r_idx, column=2).font, ws1.cell(row=r_idx,
                                                                                              column=2).border, ws1.cell(
            row=r_idx, column=2).number_format = align_right, body_font, cell_border, '#,##0'
        ws1.cell(row=r_idx, column=3).alignment, ws1.cell(row=r_idx, column=3).font, ws1.cell(row=r_idx,
                                                                                              column=3).border, ws1.cell(
            row=r_idx, column=3).number_format = align_right, body_font, cell_border, '0.00%'

    for col in range(1, 4):
        cell = ws1.cell(row=8, column=col)
        cell.font, cell.fill, cell.border, cell.alignment = total_font, total_fill, total_border, (
            align_left if col == 1 else align_right)
        if col == 2: cell.number_format = '#,##0'
        if col == 3: cell.number_format = '0.00%'

    # --- Sheet 2 Setup ---
    ws2 = wb.create_sheet(title="Chronological Trends")
    ws2.views.sheetView[0].showGridLines = True

    ws2.append([])
    ws2.append(["Table 2. Chronological Media Framing Frequency and Percentage Matrix"])
    ws2.cell(row=2, column=1).font = Font(name=font_family, size=12, bold=True)
    ws2.append([])

    ws2.append(["", "Negative", "", "Neutral", "", "Positive", "", "Annual Total", ""])
    ws2.append(["Publication Year", "f", "%", "f", "%", "f", "%", "f", "%"])
    ws2.merge_cells("B4:C4");
    ws2.merge_cells("D4:E4");
    ws2.merge_cells("F4:G4");
    ws2.merge_cells("H4:I4")

    years = sorted(df['Year'].unique())
    data_start_row = 6

    for y in years:
        row_data = [y]
        for s in ['Negative', 'Neutral', 'Positive']:
            f = int(((df['Year'] == y) & (df['RoBERTa_Full_Sentiment'] == s)).sum())
            row_data.extend([f,
                             f"=B{data_start_row}/H{data_start_row}" if s == 'Negative' else f"=D{data_start_row}/H{data_start_row}" if s == 'Neutral' else f"=F{data_start_row}/H{data_start_row}"])
        row_data.extend([f"=SUM(B{data_start_row},D{data_start_row},F{data_start_row})", f"=H{data_start_row}/$H$11"])
        ws2.append(row_data)
        data_start_row += 1

    ws2.append(["Grand Total", "=SUM(B6:B10)", "=B11/$H$11", "=SUM(D6:D10)", "=D11/$H$11", "=SUM(F6:F10)", "=F11/$H$11",
                "=SUM(H6:H10)", "=H11/$H$11"])

    # Format Sheet 2 Styles
    for r in [4, 5]:
        for col in range(1, 10):
            cell = ws2.cell(row=r, column=col)
            cell.font, cell.fill, cell.alignment = header_font, header_fill, align_center

    for r_idx in range(6, 11):
        ws2.cell(row=r_idx, column=1).alignment, ws2.cell(row=r_idx, column=1).font, ws2.cell(row=r_idx,
                                                                                              column=1).border = align_center, body_font, cell_border
        for col_idx in range(2, 10):
            cell = ws2.cell(row=r_idx, column=col_idx)
            cell.font, cell.border, cell.alignment = body_font, cell_border, align_right
            cell.number_format = '#,##0' if col_idx % 2 == 0 else '0.00%'

    for col_idx in range(1, 10):
        cell = ws2.cell(row=11, column=col_idx)
        cell.font, cell.fill, cell.border, cell.alignment = total_font, total_fill, total_border, (
            align_left if col_idx == 1 else align_right)
        if col_idx > 1: cell.number_format = '#,##0' if col_idx % 2 == 0 else '0.00%'

    # Auto-fit utility
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value and not str(cell.value).startswith("="):
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(excel_output)
    print(f"Excel Table Saved: {excel_output}")
    print("\n🎉 Pipeline Execution Successful! All charts and tables have been compiled into your project directory.")


if __name__ == '__main__':
    main()