import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import classification_report
import numpy as np


def main():
    project_dir = r"C:\Users\4Ps BARMM\Desktop\PythonProject3"
    file_path = os.path.join(project_dir, "blind human annotation Results for coding.xlsx")

    print("📦 Reading the source Excel file...")
    df = pd.read_excel(file_path)

    # Clean column headers by stripping accidental whitespaces
    df.columns = df.columns.str.strip()

    # Drop rows that are completely blank in our critical tracking columns
    df = df.dropna(subset=['Final Sentiment', 'vader_prediction', 'roberta_prediction'])

    # =========================================================================
    # 🧹 FORGIVING CLEANING & MAPPING STEP (Ensures all 343 rows match)
    # =========================================================================
    # Convert the human column to uppercase strings and remove hidden spaces
    df['Final Sentiment'] = df['Final Sentiment'].astype(str).str.strip().str.upper()

    # Expanded dictionary to capture any casing variants safely
    human_map = {
        'P': 'positive', 'POSITIVE': 'positive',
        'NU': 'neutral', 'NEUTRAL': 'neutral',
        'N': 'negative', 'NEGATIVE': 'negative'
    }
    df['Human_Clean'] = df['Final Sentiment'].map(human_map)

    # Standardize model prediction columns to lowercase strings
    df['vader_clean'] = df['vader_prediction'].astype(str).str.strip().str.lower()
    df['roberta_clean'] = df['roberta_prediction'].astype(str).str.strip().str.lower()

    # Retain only rows containing valid structural target labels
    valid_categories = ['negative', 'neutral', 'positive']
    df = df[df['Human_Clean'].isin(valid_categories) &
            df['vader_clean'].isin(valid_categories) &
            df['roberta_clean'].isin(valid_categories)]

    total_rows = len(df)
    print(f"📋 Total Rows Successfully Processed: {total_rows}")
    # =========================================================================

    # 📊 Compute statistical metrics using scikit-learn
    vader_report = classification_report(df['Human_Clean'], df['vader_clean'], labels=valid_categories,
                                         output_dict=True)
    roberta_report = classification_report(df['Human_Clean'], df['roberta_clean'], labels=valid_categories,
                                           output_dict=True)

    # =========================================================================
    # 💻 PART 1: GENERATING THE REVISED EXCEL STYLED TABLES
    # =========================================================================
    print("📊 Organizing and styling evaluation tables directly into Excel...")
    wb = openpyxl.Workbook()

    font_name = "Calibri"
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    total_fill = PatternFill(start_color="E4EDF6", end_color="E4EDF6", fill_type="solid")
    border_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    double_border = Border(left=border_side, right=border_side, top=border_side,
                           bottom=Side(border_style="double", color="111827"))

    # Sheet 1: VADER
    ws_vader = wb.active
    ws_vader.title = "VADER Evaluation"
    ws_vader.append(["VADER Lexicon-Based Sentiment Evaluation Matrix"])
    ws_vader.append(["Performance metrics compared against the Blind Human Expert Ground Truth"])
    ws_vader.append([])
    ws_vader.append(["Sentiment Class", "Precision", "Recall", "F1-Score", "Support (Sample Count)"])

    for cls in valid_categories:
        metrics = vader_report[cls]
        ws_vader.append(
            [cls.capitalize(), metrics['precision'], metrics['recall'], metrics['f1-score'], metrics['support']])
    ws_vader.append(["Macro Average", "", "", vader_report['macro avg']['f1-score'], vader_report['macro avg']['support']])

    # Sheet 2: RoBERTa
    ws_roberta = wb.create_sheet(title="RoBERTa Evaluation")
    ws_roberta.append(["RoBERTa Transformer-Based Sentiment Evaluation Matrix"])
    ws_roberta.append(["Performance metrics compared against the Blind Human Expert Ground Truth"])
    ws_roberta.append([])
    ws_roberta.append(["Sentiment Class", "Precision", "Recall", "F1-Score", "Support (Sample Count)"])

    for cls in valid_categories:
        metrics = roberta_report[cls]
        ws_roberta.append(
            [cls.capitalize(), metrics['precision'], metrics['recall'], metrics['f1-score'], metrics['support']])
    ws_roberta.append(["Macro Average", "", "", roberta_report['macro avg']['f1-score'], roberta_report['macro avg']['support']])

    # Apply formatting patterns
    for ws in [ws_vader, ws_roberta]:
        ws.cell(row=1, column=1).font = Font(name=font_name, size=14, bold=True, color="1F4E79")
        ws.cell(row=2, column=1).font = Font(name=font_name, size=10, italic=True, color="595959")

        for col in range(1, 6):
            cell = ws.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in range(5, 8):
            row_fill = zebra_fill if r % 2 == 0 else white_fill
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
            for col in range(2, 5):
                ws.cell(row=r, column=col).number_format = "0.00"  # PINALITAN: Mula 0.0000 ginawang 2 decimal places
                ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=5).number_format = "#,##0"
            ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = row_fill
                ws.cell(row=r, column=col).border = thin_border

        last_row = 8
        ws.cell(row=last_row, column=1).font = Font(bold=True)
        ws.cell(row=last_row, column=4).font = Font(bold=True, color="1F4E79")
        ws.cell(row=last_row, column=4).number_format = "0.00"  # PINALITAN: Dalawang decimal places para sa Macro Average
        ws.cell(row=last_row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=last_row, column=5).font = Font(bold=True)
        ws.cell(row=last_row, column=5).number_format = "#,##0"
        ws.cell(row=last_row, column=5).alignment = Alignment(horizontal="center")
        for col in range(1, 6):
            ws.cell(row=last_row, column=col).fill = total_fill
            ws.cell(row=last_row, column=col).border = double_border

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 22

    output_excel_path = os.path.join(project_dir, "Sentiment_Model_Evaluation_Results Revised for hardbound.xlsx")
    wb.save(output_excel_path)
    print("💾 Spreadsheet generated successfully.")

    # =========================================================================
    # 📈 PART 2: GENERATING THE REVISED INTERSECTION CHART (BOTH vs NEITHER)
    # =========================================================================
    print("📈 Extracting alignments and plotting revised alignment figure...")

    # Compute Intersections
    both_match = df[(df['vader_clean'] == df['Human_Clean']) & (df['roberta_clean'] == df['Human_Clean'])]
    none_match = df[(df['vader_clean'] != df['Human_Clean']) & (df['roberta_clean'] != df['Human_Clean'])]

    # Define categories and data
    categories_revised = ['Both Models\nMatch Human', 'Neither Model\nMatches Human']
    counts_revised = [len(both_match), len(none_match)]
    percentages_revised = [(count / total_rows) * 100 for count in counts_revised]

    # Plotting
    plt.figure(figsize=(7, 6))
    sns.set_theme(style="white")
    colors_revised = ['#2E5597', '#C00000']  # Asul para sa Both, Pula para sa Neither

    ax_revised = sns.barplot(x=categories_revised, y=percentages_revised, hue=categories_revised,
                             palette=colors_revised, legend=False)

    # Annotations
    for i, p in enumerate(ax_revised.patches):
        ax_revised.annotate(f"{p.get_height():.2f}%\n({counts_revised[i]} rows)",
                            (p.get_x() + p.get_width() / 2., p.get_height()),
                            ha='center', va='center', xytext=(0, 12), textcoords='offset points',
                            fontsize=11, fontweight='bold')

    plt.title('Comparison of Perfect Alignment vs. Total Model Failure', fontsize=13, fontweight='bold', pad=20)
    plt.ylabel('Percentage of Validated Sample (%)', fontsize=11, fontweight='bold')
    plt.ylim(0, max(percentages_revised) + 15)
    plt.tight_layout()

    # Save and Show
    plt.savefig(os.path.join(project_dir, "both_vs_neither_alignment2.png"), dpi=300)
    plt.show()
    # =========================================================================
    # 📈 CHART 2: Overall Model Performance (RoBERTa Total vs VADER Total)
    # =========================================================================
    roberta_total = df[df['roberta_clean'] == df['Human_Clean']]
    vader_total = df[df['vader_clean'] == df['Human_Clean']]

    categories2 = ['Total RoBERTa\nMatches Human', 'Total VADER\nMatches Human']
    counts2 = [len(roberta_total), len(vader_total)]
    percentages2 = [(count / total_rows) * 100 for count in counts2]

    plt.figure(figsize=(7, 6))
    colors2 = ['#4F81BD', '#ED7D31']

    ax2 = sns.barplot(x=categories2, y=percentages2, hue=categories2, palette=colors2, legend=False)

    for i, p in enumerate(ax2.patches):
        ax2.annotate(f"{p.get_height():.2f}%\n({counts2[i]} rows)",
                     (p.get_x() + p.get_width() / 2., p.get_height()),
                     ha='center', va='center', xytext=(0, 12), textcoords='offset points',
                     fontsize=10, fontweight='bold')

    # Fix para sa cut-off title gamit ang \n at rect adjustment
    plt.title('Comparative Alignment Rates of RoBERTa and VADER\nwith Human Ground Truth',
              fontsize=13, fontweight='bold', pad=25)
    plt.ylabel('Percentage of Validated Sample (%)', fontsize=11, fontweight='bold')
    plt.ylim(0, max(percentages2) + 15)

    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    plt.savefig(os.path.join(project_dir, "overall_model_performance_Final2.png"), dpi=300)
    plt.show()

    # =========================================================================
    # 📈 CHART 3: Mutual Error Breakdown Chart
    # =========================================================================
    if len(none_match) > 0:
        plt.figure(figsize=(7, 5))
        error_counts = none_match['Human_Clean'].value_counts().reindex(valid_categories, fill_value=0)
        x_labels = [c.capitalize() for c in error_counts.index]

        ax3 = sns.barplot(x=x_labels, y=error_counts.values, hue=x_labels, palette='Reds_r', legend=False)

        for p in ax3.patches:
            ax3.annotate(f"{int(p.get_height())} headlines",
                         (p.get_x() + p.get_width() / 2., p.get_height()),
                         ha='center', va='center', xytext=(0, 8), textcoords='offset points',
                         fontsize=10, fontweight='bold')

        plt.title('True Sentiment Distribution of Total Conflicts\n(When Both VADER and RoBERTa Fail)', fontsize=12,
                  fontweight='bold', pad=15)
        plt.xlabel('Actual Human Sentiment Category', fontsize=10, fontweight='bold')
        plt.ylabel('Number of Misclassified Headlines', fontsize=10, fontweight='bold')
        plt.ylim(0, max(error_counts.values) + (max(error_counts.values) * 0.2 if max(error_counts.values) > 0 else 5))
        plt.tight_layout()

        chart3_path = os.path.join(project_dir, "mutual_error_breakdown.png")
        plt.savefig(chart3_path, dpi=300)
        plt.show()
        # =========================================================================
        # 📈 PART 3: COMPARATIVE F1-SCORE CHART (RoBERTa vs VADER)
        # =========================================================================
        print("📈 Generating Comparative F1-Score Chart...")

        classes = [c.capitalize() for c in valid_categories]
        roberta_f1 = [roberta_report[c]['f1-score'] for c in valid_categories]
        vader_f1 = [vader_report[c]['f1-score'] for c in valid_categories]

        x = np.arange(len(classes))
        width = 0.35

        plt.figure(figsize=(9, 6))
        rects1 = plt.bar(x - width / 2, roberta_f1, width, label='RoBERTa', color='#1F4E79')
        rects2 = plt.bar(x + width / 2, vader_f1, width, label='VADER', color='#E46C0A')

        plt.ylabel('F1-Score', fontsize=11, fontweight='bold')
        plt.title('Comparative F1-Score Performance: RoBERTa vs VADER', fontsize=13, fontweight='bold', pad=20)
        plt.xticks(x, classes)
        plt.ylim(0, 1.05)
        plt.legend()

        # Function to label bars
        def autolabel(rects):
            for rect in rects:
                height = rect.get_height()
                plt.annotate(f'{height:.2f}',  # PINALITAN: Binago mula :.3f ginawang :.2f para sa labels ng bar
                             xy=(rect.get_x() + rect.get_width() / 2, height),
                             xytext=(0, 3),
                             textcoords="offset points",
                             ha='center', va='bottom', fontsize=10, fontweight='bold')

        autolabel(rects1)
        autolabel(rects2)

        plt.tight_layout()
        chart3_path = os.path.join(project_dir, "Comparative_F1_Performance2.png")
        plt.savefig(chart3_path, dpi=300)
        plt.show()

    print("\n" + "=" * 50)
    print("🎉 ALL TASKS COMPLETED SUCCESSFULLY!")
    print("=" * 50)
    print(f"📊 Rows processed: {total_rows}/343")
    print(f"🗂️  Excel workbook saved:  'Sentiment_Model_Evaluation_ResultsFinal.xlsx'") # Inayos ang pangalan ng file dito
    print(f"🖼️  Chart 1 saved:         'model_human_alignment_intersection5.png'")
    print(f"🖼️  Chart 2 saved:         'mutual_error_breakdown.png' (if rows existed)")
    print(f"🖼️  Chart 3 saved:         'Comparative_F1_Performance2.png'")
    print("=" * 50)


if __name__ == '__main__':
    main()